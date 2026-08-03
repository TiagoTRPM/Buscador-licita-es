# main.py
import io
import os
import re
from datetime import datetime, timedelta
from fastapi.concurrency import run_in_threadpool

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from config import ESTADOS, BASE_DIR, MODALIDADES_DISPONIVEIS
from pncp_buscador import buscar_licitacoes_eficiente
from pncp_extrator import extrair_licitacoes
from expansor import expandir_termo
from etp_ia import sintetizar_campos_etp
import cache as cache_module
import atualizador
import etp_manager
from etp_gerador import gerar_word
from etp_importador import extrair_campos_etp, garantir_docx

# =====================================================
# INICIALIZAÇÃO
# =====================================================

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app):
    atualizador.iniciar_monitor_noturno()
    yield

app = FastAPI(lifespan=lifespan)


app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

_ultima_busca: dict | None = None

def _data_padrao_inicial():
    return (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d")

def _data_padrao_final():
    return datetime.now().strftime("%Y-%m-%d")



# =====================================================
# ROTAS: ATUALIZADOR DE CACHE
# =====================================================

@app.post("/api/atualizar/iniciar")
async def atualizar_iniciar():
    if not _ultima_busca:
        return {"ok": False, "erro": "Faça uma busca primeiro"}
    lics     = _ultima_busca.get("licitacoes", [])
    pendentes = cache_module.listar_pendentes(lics)
    return atualizador.iniciar(pendentes)


@app.post("/api/atualizar/cancelar")
async def atualizar_cancelar():
    atualizador.cancelar()
    return {"ok": True}


@app.get("/api/atualizar/progresso")
async def atualizar_progresso():
    return atualizador.progresso()


@app.get("/api/cache/stats")
async def cache_stats():
    return cache_module.stats()


@app.get("/api/valor/{numero_controle:path}")
async def api_valor_cache(numero_controle: str):
    """Tenta cache primeiro, depois busca na API."""
    cached = cache_module.get(numero_controle)
    if cached and cached.get("valor"):
        return {"valor": cached["valor"], "fonte": "cache"}

    import requests as req
    import time as tm
    try:
        partes    = numero_controle.split("/")
        ano       = partes[1].strip()
        segmentos = partes[0].split("-")
        cnpj      = segmentos[0].strip()
        seq       = str(int(segmentos[2].strip()))
    except Exception:
        return {"valor": None, "erro": "Número de controle inválido"}

    url = f"https://pncp.gov.br/api/consulta/v1/contratacoes/{cnpj}/{ano}/{seq}"
    try:
        tm.sleep(0.5)
        resp = req.get(url, headers={
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36",
            "Referer": "https://pncp.gov.br/app/editais",
        }, timeout=10)
        if resp.status_code == 200:
            d = resp.json()
            valor = d.get("valorTotalEstimado") or d.get("valorTotalHomologado")
            valor_fmt = None
            if valor:
                valor_fmt = (
                    f"R$ {float(valor):,.2f}"
                    .replace(",", "X")
                    .replace(".", ",")
                    .replace("X", ".")
                )
            unidade = d.get("unidadeOrgao") or {}
            orgao   = d.get("orgaoEntidade") or {}
            registro = {
                "valor":          valor_fmt,
                "objeto":         (d.get("objetoCompra") or "")[:200],
                "orgao":          orgao.get("razaoSocial") or "",
                "municipio":      unidade.get("municipioNome") or "",
                "uf":             unidade.get("ufSigla") or "",
                "modalidade":     d.get("modalidadeNome") or "",
                "numero_edital":  d.get("numeroCompra") or numero_controle,
                "data_publicacao": (d.get("dataPublicacaoPncp") or "")[:10],
                "situacao":       d.get("situacaoCompraNome") or "",
            }
            cache_module.set(numero_controle, registro)
            return {"valor": valor_fmt, "fonte": "api"}
        return {"valor": None, "erro": f"API retornou {resp.status_code}"}
    except Exception as e:
        return {"valor": None, "erro": str(e)}


def _formatar_brl(valor):
    if valor in (None, "", 0):
        return None
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return None


@app.get("/api/itens/{numero_controle:path}")
async def api_itens_contratacao(numero_controle: str):
    """Consulta e guarda os itens, incluindo valores unitários, da contratação no PNCP."""
    cached = cache_module.get(numero_controle) or {}
    if cached.get("itens"):
        return {"itens": cached["itens"], "fonte": "cache"}
    try:
        partes = numero_controle.split("/")
        ano = partes[1].strip()
        segmentos = partes[0].split("-")
        cnpj, seq = segmentos[0].strip(), str(int(segmentos[2].strip()))
    except Exception:
        return {"itens": [], "erro": "Número de controle inválido"}

    import requests as req
    url = f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj}/compras/{ano}/{seq}/itens"
    try:
        resposta = req.get(url, headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0", "Referer": "https://pncp.gov.br/app/editais"}, timeout=15)
        if resposta.status_code != 200:
            return {"itens": [], "erro": f"API retornou {resposta.status_code}"}
        bruto = resposta.json()
        if isinstance(bruto, list):
            lista = bruto
        elif isinstance(bruto, dict):
            lista = bruto.get("data") or bruto.get("items") or []
        else:
            lista = []
        itens = []
        for indice, item in enumerate(lista, 1):
            unitario = item.get("valorUnitarioEstimado") or item.get("valorUnitarioHomologado") or item.get("valorUnitario")
            total = item.get("valorTotal") or item.get("valorTotalEstimado") or item.get("valorTotalHomologado")
            itens.append({
                "id": str(item.get("numeroItem") or item.get("sequencialItem") or indice),
                "descricao": item.get("descricao") or item.get("descricaoItem") or "Item sem descrição",
                "quantidade": item.get("quantidade") or item.get("quantidadeItem") or "",
                "unidade": item.get("unidadeMedida") or item.get("unidade") or "",
                "valor_unitario": _formatar_brl(unitario),
                "valor_total": _formatar_brl(total),
            })
        cache_module.set(numero_controle, {"itens": itens})
        return {"itens": itens, "fonte": "api"}
    except Exception as e:
        return {"itens": [], "erro": str(e)}



# =====================================================
# ROTAS: ETP
# =====================================================

@app.get("/etp", response_class=HTMLResponse)
async def etp_lista(request: Request):
    etps = etp_manager.listar_etps()
    return templates.TemplateResponse(
        request=request,
        name="etp.html",
        context={"etps": etps},
    )


@app.post("/etp/novo")
async def etp_novo(request: Request):
    form = await request.form()
    titulo = form.get("titulo", "Novo ETP").strip() or "Novo ETP"
    etp = etp_manager.novo_etp(titulo)
    return RedirectResponse(f"/etp/{etp['id']}", status_code=303)


@app.post("/api/etp/novo")
async def api_etp_novo(request: Request):
    dados = await request.json()
    titulo = (dados.get("titulo") or "Novo ETP").strip() or "Novo ETP"
    etp = etp_manager.novo_etp(titulo)
    return {"ok": True, "id": etp["id"], "titulo": etp["titulo"]}


@app.get("/etp/{etp_id}", response_class=HTMLResponse)
async def etp_form(request: Request, etp_id: str):
    etp = etp_manager.get_etp(etp_id)
    if not etp:
        return RedirectResponse("/etp", status_code=303)
    return templates.TemplateResponse(
        request=request,
        name="etp_form.html",
        context={"etp": etp},
    )


@app.post("/etp/{etp_id}/salvar")
async def etp_salvar(request: Request, etp_id: str):
    form = await request.form()

    cabecalho = {
        "orgao":       form.get("orgao", ""),
        "municipio":   form.get("municipio", ""),
        "estado":      form.get("estado", ""),
        "processo":    form.get("processo", ""),
        "responsavel": form.get("responsavel", ""),
        "cargo":       form.get("cargo", ""),
        "data":        form.get("data", ""),
        "modo_cabecalho": form.get("modo_cabecalho", "dinamico"),
        "brasao_path": form.get("brasao_path", ""),
    }

    campos_fixos = {
        "descricao_necessidade":      form.get("descricao_necessidade", ""),
        "previsao_pca":               form.get("previsao_pca", ""),
        "requisitos_contratacao":     form.get("requisitos_contratacao", ""),
        "descricao_solucao":          form.get("descricao_solucao", ""),
        "estimativa_valor":           form.get("estimativa_valor", ""),
        "justificativa_parcelamento": form.get("justificativa_parcelamento", ""),
        "resultados_pretendidos":     form.get("resultados_pretendidos", ""),
        "providencias_previas":       form.get("providencias_previas", ""),
        "contratacoes_correlatas":    form.get("contratacoes_correlatas", ""),
        "impactos_ambientais":        form.get("impactos_ambientais", ""),
        "posicionamento_conclusivo":  form.get("posicionamento_conclusivo", ""),
    }

    etapas_puladas = [e for e in form.get("etapas_puladas", "").split(",") if e]

    titulo = form.get("titulo", "")

    etp_manager.salvar_etp(etp_id, {
        "titulo":       titulo,
        "cabecalho":    cabecalho,
        "campos_fixos": campos_fixos,
        "etapas_puladas": etapas_puladas,
    })

    return RedirectResponse(f"/etp/{etp_id}", status_code=303)


@app.post("/etp/{etp_id}/importar")
async def etp_importar(etp_id: str, arquivo: UploadFile = File(...)):
    if not etp_manager.get_etp(etp_id):
        return RedirectResponse("/etp", status_code=303)

    if not (arquivo.filename or "").lower().endswith((".docx", ".doc", ".pdf", ".odt")):
        return RedirectResponse(f"/etp/{etp_id}?erro_importacao=formato", status_code=303)

    conteudo = await arquivo.read()
    if not conteudo or len(conteudo) > 10 * 1024 * 1024:
        return RedirectResponse(f"/etp/{etp_id}?erro_importacao=arquivo", status_code=303)

    try:
        conteudo_docx = await run_in_threadpool(garantir_docx, arquivo.filename, conteudo)
        campos = extrair_campos_etp(conteudo_docx)
    except Exception as e:
        print(f"[etp] Erro na importacao: {e}")
        return RedirectResponse(f"/etp/{etp_id}?erro_importacao=leitura", status_code=303)

    if not campos:
        return RedirectResponse(f"/etp/{etp_id}?erro_importacao=blocos", status_code=303)

    etp_manager.importar_campos_fixos(etp_id, campos)
    return RedirectResponse(f"/etp/{etp_id}?importados={len(campos)}#campos", status_code=303)


@app.post("/etp/{etp_id}/sintetizar")
async def etp_sintetizar(
    etp_id: str,
    arquivos: list[UploadFile] = File(...),
    base_index: int = Form(...),
    objeto: str = Form(...),
):
    if not etp_manager.get_etp(etp_id):
        return RedirectResponse("/etp", status_code=303)
    if not objeto.strip() or not 3 <= len(arquivos) <= 5 or not 0 <= base_index < len(arquivos):
        return RedirectResponse(f"/etp/{etp_id}?erro_sintese=quantidade", status_code=303)
    if any(not (arquivo.filename or "").lower().endswith((".docx", ".doc", ".pdf", ".odt")) for arquivo in arquivos):
        return RedirectResponse(f"/etp/{etp_id}?erro_sintese=formato", status_code=303)

    try:
        documentos = []
        for arquivo in arquivos:
            conteudo = await arquivo.read()
            if not conteudo or len(conteudo) > 10 * 1024 * 1024:
                raise ValueError("Arquivo inválido")
            
            conteudo_docx = await run_in_threadpool(garantir_docx, arquivo.filename, conteudo)
            documentos.append(extrair_campos_etp(conteudo_docx))
            
        if any(not documento for documento in documentos):
            raise ValueError("Nenhum bloco identificado")

        campos = await run_in_threadpool(
            sintetizar_campos_etp,
            documentos[base_index],
            [documento for indice, documento in enumerate(documentos) if indice != base_index],
            objeto.strip(),
            etp_manager.get_etp(etp_id).get("itens", []),
        )
    except Exception as erro:
        print(f"[etp] Erro ao sintetizar ETPs: {erro}")
        # Redirecionar incluindo a mensagem de erro para depuração
        import urllib.parse
        erro_msg = urllib.parse.quote(str(erro))
        return RedirectResponse(f"/etp/{etp_id}?erro_sintese=processamento&detalhe={erro_msg}", status_code=303)

    if not campos:
        return RedirectResponse(f"/etp/{etp_id}?erro_sintese=resultado", status_code=303)
    etp_manager.importar_campos_fixos(etp_id, campos)
    return RedirectResponse(f"/etp/{etp_id}?sintetizados={len(campos)}#campos", status_code=303)


@app.post("/etp/{etp_id}/item/add")
async def etp_item_add(request: Request, etp_id: str):
    form = await request.form()
    item = {
        "nome":        " ".join(form.get("nome", "").split())[:100],
        "descricao":   form.get("descricao", "").strip(),
        "quantidade":  form.get("quantidade", "").strip(),
        "unidade":     form.get("unidade", "un").strip() or "un",
    }
    if not item["nome"] or not item["quantidade"]:
        return RedirectResponse(f"/etp/{etp_id}#itens", status_code=303)

    etp = etp_manager.adicionar_item(etp_id, item)
    if not etp:
        return RedirectResponse("/etp", status_code=303)
    return RedirectResponse(f"/etp/{etp_id}#itens", status_code=303)


@app.post("/etp/{etp_id}/itens/incluir-levantamento")
async def etp_itens_incluir_levantamento(etp_id: str):
    etp_manager.adicionar_todos_itens_do_levantamento(etp_id)
    return RedirectResponse(f"/etp/{etp_id}#itens", status_code=303)


@app.post("/etp/{etp_id}/item/{item_id}/remover")
async def etp_item_remover(etp_id: str, item_id: str):
    etp_manager.remover_item(etp_id, item_id)
    return RedirectResponse(f"/etp/{etp_id}#itens", status_code=303)


@app.post("/etp/{etp_id}/item/{item_id}/editar")
async def etp_item_editar(request: Request, etp_id: str, item_id: str):
    form = await request.form()
    nome = form.get("nome", "").strip()
    if nome:
        etp_manager.editar_nome_item(etp_id, item_id, nome)
    return RedirectResponse(f"/etp/{etp_id}#itens", status_code=303)


@app.post("/etp/{etp_id}/campo-livre/add")
async def etp_campo_livre_add(request: Request, etp_id: str):
    form = await request.form()
    etp_manager.adicionar_campo_livre(
        etp_id,
        nome=form.get("nome", ""),
        conteudo=form.get("conteudo", ""),
    )
    return RedirectResponse(f"/etp/{etp_id}#campos-livres", status_code=303)


@app.get("/etp/{etp_id}/exportar")
async def etp_exportar(request: Request, etp_id: str):
    etp = etp_manager.get_etp(etp_id)
    if not etp:
        return RedirectResponse("/etp", status_code=303)

    obrigatorios = {
        "descricao_necessidade": "Descrição da necessidade",
        "estimativa_valor": "Estimativa do valor",
        "descricao_solucao": "Descrição da solução",
        "justificativa_parcelamento": "Justificativa do parcelamento",
        "resultados_pretendidos": "Resultados pretendidos",
        "providencias_previas": "Providências prévias",
        "contratacoes_correlatas": "Contratações correlatas",
        "impactos_ambientais": "Impactos ambientais",
        "posicionamento_conclusivo": "Posicionamento conclusivo sobre a viabilidade",
    }
    faltantes = [nome for campo, nome in obrigatorios.items() if not etp.get("campos_fixos", {}).get(campo, "").strip()]
    if not etp.get("itens"):
        faltantes.append("Estimativa de quantidades (inclua ao menos um item)")
    if not etp.get("levantamento"):
        faltantes.append("Levantamento de mercado (inclua ao menos uma pesquisa ou referência)")
    if faltantes:
        return templates.TemplateResponse(
            request=request, name="etp_form.html",
            context={"etp": etp, "erros_exportacao": faltantes}, status_code=422,
        )

    try:
        docx_bytes = await run_in_threadpool(gerar_word, etp)
        titulo = etp.get("titulo", "etp").replace(" ", "_")
        return StreamingResponse(
            io.BytesIO(docx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename=ETP_{titulo}.docx"},
        )
    except ImportError as e:
        return HTMLResponse(f"Erro: {e}", status_code=500)


@app.delete("/etp/{etp_id}")
async def etp_deletar(etp_id: str):
    etp_manager.deletar_etp(etp_id)
    return {"ok": True}


@app.post("/etp/{etp_id}/levantamento")
async def etp_salvar_levantamento(request: Request, etp_id: str):
    dados = await request.json()
    termo = dados.get("termo", "")
    referencias = dados.get("referencias", [])
    etp_manager.salvar_levantamento(etp_id, termo, referencias)
    return {"ok": True}


@app.post("/api/etp/{etp_id}/referencias/adicionar")
async def etp_adicionar_referencia(etp_id: str, request: Request):
    """Adiciona ao ETP uma licitação presente na consulta atual."""
    if not etp_manager.get_etp(etp_id):
        return {"ok": False, "erro": "ETP não encontrado"}

    dados = await request.json()
    numero_controle = dados.get("numero_controle", "")
    termo = dados.get("termo", "")
    item_id = str(dados.get("item_id", ""))
    if not _ultima_busca or not numero_controle:
        return {"ok": False, "erro": "Resultado da consulta não encontrado"}

    licitacao = next(
        (l for l in _ultima_busca.get("licitacoes", [])
         if l.get("_numeroControlePNCP") == numero_controle),
        None,
    )
    if not licitacao:
        return {"ok": False, "erro": "Esta licitação não está mais na consulta atual"}

    cached = cache_module.get(numero_controle)
    item = next((i for i in (cached or {}).get("itens", []) if str(i.get("id")) == item_id), None)
    referencia = {
        "orgao": licitacao.get("orgao", ""),
        "municipio": licitacao.get("municipio", ""),
        "uf": licitacao.get("uf", ""),
        "data_publicacao": licitacao.get("publicacao", ""),
        "objeto": item.get("descricao") if item else licitacao.get("objeto", ""),
        "modalidade": licitacao.get("modalidade", ""),
        "link": licitacao.get("link", ""),
        "numero_controle": numero_controle,
        "item_id": item_id or None,
        "descricao_item": item.get("descricao", "") if item else "",
        "quantidade": item.get("quantidade", "") if item else "",
        "unidade": item.get("unidade", "") if item else "",
        "valor_unitario": item.get("valor_unitario") if item else None,
        "valor": item.get("valor_total") if item else (cached.get("valor") if cached else licitacao.get("valor", "")),
    }
    _, adicionada = etp_manager.adicionar_referencia(etp_id, termo, referencia)
    return {"ok": True, "adicionada": adicionada}



# =====================================================
# ROTA: API — referências de preço para ETP
# =====================================================

@app.get("/api/etp/referencias")
async def api_etp_referencias(termo: str = ""):
    """
    Retorna os resultados da última busca formatados para o levantamento de mercado do ETP.
    Extrai valor unitário dos itens quando disponível no cache.
    """
    if not _ultima_busca:
        return {"referencias": []}

    licitacoes = _ultima_busca.get("licitacoes", [])
    termo_lower = termo.lower()

    referencias = []
    for l in licitacoes:
        # Filtra pelo termo no objeto
        if termo_lower and termo_lower not in l.get("objeto", "").lower():
            continue

        nc = l.get("_numeroControlePNCP", "")

        # Tenta pegar do cache
        cached = cache_module.get(nc) if nc else None

        ref = {
            "orgao":           l.get("orgao", ""),
            "municipio":       l.get("municipio", ""),
            "uf":              l.get("uf", ""),
            "data_publicacao": l.get("publicacao", "")[:10] if l.get("publicacao") else "",
            "objeto":          l.get("objeto", "")[:150],
            "modalidade":      l.get("modalidade", ""),
            "link":            l.get("link", ""),
            "numero_controle": nc,
            "valor_unitario":  cached.get("valor") if cached else None,
            "valor":           cached.get("valor") if cached else l.get("valor", ""),
        }
        referencias.append(ref)

    # Limita a 20 referências mais recentes
    return {"referencias": referencias[:20], "total": len(referencias)}


# =====================================================
# ROTA: INDEX
# =====================================================

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    response = templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "estados":      ESTADOS,
            "modalidades":  MODALIDADES_DISPONIVEIS,
            "data_inicial": _data_padrao_inicial(),
            "data_final":   _data_padrao_final(),
        },
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


@app.get("/buscar")
async def buscar_get_invalido():
    return RedirectResponse("/", status_code=303)


# =====================================================
# ROTA: API — expandir termos via Gemini
# =====================================================

@app.get("/api/expandir")
async def api_expandir(termo: str = ""):
    if not termo or len(termo.strip()) < 2:
        return {"termos_incluir": [], "termos_excluir": []}
    resultado = await run_in_threadpool(expandir_termo, termo.strip())
    return resultado


# =====================================================
# ROTA: API — municípios de um estado
# =====================================================

@app.get("/api/municipios/{uf}")
async def api_municipios(uf: str, termo: str = "", data_inicial: str = "", data_final: str = ""):
    if not termo:
        return {"municipios": []}

    hoje = datetime.now().date()
    try:
        dt_ini = datetime.strptime(data_inicial, "%Y-%m-%d").date() if data_inicial else hoje - timedelta(days=180)
        dt_fim = datetime.strptime(data_final,   "%Y-%m-%d").date() if data_final   else hoje
    except ValueError:
        dt_ini = hoje - timedelta(days=180)
        dt_fim = hoje

    dias = max((hoje - dt_ini).days + (dt_fim - hoje).days, 1)
    dados_json = buscar_licitacoes_eficiente(termo, ufs=[uf], limite=500, dias=dias)
    licitacoes = extrair_licitacoes(dados_json)
    municipios = sorted({l["municipio"] for l in licitacoes})
    return {"municipios": municipios}


# =====================================================
# ROTA: BUSCA COM SSE (progresso em tempo real)
# =====================================================

@app.post("/buscar-stream")
async def buscar_stream(request: Request):

    form = await request.form()

    termo               = form.get("termo", "").strip()
    data_inicial        = form.get("data_inicial", "")
    data_final          = form.get("data_final", "")
    limite_registros    = int(form.get("limite_registros", 200))
    incluir_sem_periodo = form.get("incluir_sem_periodo", "nao")
    ufs_selecionadas    = form.getlist("ufs") or ["PR"]
    filtros_cidade      = form.getlist("cidades_filtro")
    mods_form           = form.getlist("modalidades")
    modalidades_selecionadas = [int(m) for m in mods_form] if mods_form else list(MODALIDADES_DISPONIVEIS.keys())

    termos_extras  = form.getlist("termos_incluir")
    termos_excluir = form.getlist("termos_excluir")

    if not termo:
        return RedirectResponse("/", status_code=303)

    hoje = datetime.now().date()
    try:
        dt_ini = datetime.strptime(data_inicial, "%Y-%m-%d").date() if data_inicial else hoje - timedelta(days=180)
        dt_fim = datetime.strptime(data_final,   "%Y-%m-%d").date() if data_final   else hoje
    except ValueError:
        dt_ini = hoje - timedelta(days=180)
        dt_fim = hoje

    dias = max((hoje - dt_ini).days + (dt_fim - hoje).days, 1)

    todos_termos = [termo] + [t for t in termos_extras if t.strip()]

    STOPWORDS = {
        "contratacao", "contratação", "servico", "serviço", "servicos", "serviços",
        "de", "da", "do", "das", "dos", "para", "com", "a", "e", "em", "um", "uma", "por"
    }

    async def gerar_eventos():
        todas_licitacoes = []
        vistos_link = set()

        for i, t in enumerate(todos_termos, 1):
            # Evento de progresso
            yield f"data: {{'\"tipo\": \"progresso\", \"atual\": {i}, \"total\": {len(todos_termos)}, \"termo\": \"{t}\"}}\n\n"

            dados_json = await run_in_threadpool(
                buscar_licitacoes_eficiente,
                t,
                ufs=ufs_selecionadas,
                limite=limite_registros,
                dias=dias,
                modalidades=modalidades_selecionadas,
            )

            licitacoes = extrair_licitacoes(dados_json) if dados_json else []

            # Filtro de termo no objeto
            palavras_usuario = [
                p.lower().strip() for p in t.split()
                if len(p) > 2 and p.lower() not in STOPWORDS
            ]

            novas = 0
            for l in licitacoes:
                if l["link"] in vistos_link:
                    continue

                objeto_original = l["objeto"].lower()
                objeto_limpo = re.sub(r'^\s*\[.*?\]\s*-?\s*', '', objeto_original)

                if palavras_usuario:
                    if not any(p in objeto_limpo or p in objeto_original for p in palavras_usuario):
                        continue

                vistos_link.add(l["link"])
                todas_licitacoes.append(l)
                novas += 1

            yield f"data: {{\"tipo\": \"parcial\", \"termo\": \"{t}\", \"encontrados\": {novas}, \"total_ate_agora\": {len(todas_licitacoes)}}}\n\n"

        # Filtro de exclusão
        if termos_excluir:
            excluir_lower = [t.lower() for t in termos_excluir]
            todas_licitacoes = [
                l for l in todas_licitacoes
                if not any(ex in l["objeto"].lower() for ex in excluir_lower)
            ]

        # Filtro de período
        def dentro_do_periodo(l):
            if l["_dt_ref"] is None:
                return incluir_sem_periodo == "sim"
            return dt_ini <= l["_dt_ref"] <= dt_fim

        todas_licitacoes = [l for l in todas_licitacoes if dentro_do_periodo(l)]

        # Filtro de cidades
        if filtros_cidade:
            cidades_set = set()
            for fc in filtros_cidade:
                partes = fc.split(":", 1)
                if len(partes) == 2:
                    cidades_set.add(partes[1].strip().lower())
            todas_licitacoes = [l for l in todas_licitacoes if l["municipio"].lower() in cidades_set]

        if limite_registros > 0:
            todas_licitacoes = todas_licitacoes[:limite_registros]

        # Separa por status
        abertos    = sorted([l for l in todas_licitacoes if l["status"] == "Em Aberto"],           key=lambda x: x["dias"])
        aguardando = sorted([l for l in todas_licitacoes if l["status"] == "Aguardando Abertura"], key=lambda x: x["dias"])
        julgamento = sorted([l for l in todas_licitacoes if l["status"] == "Em Julgamento"],       key=lambda x: x["dias"])
        encerrados = sorted([l for l in todas_licitacoes if l["status"] == "Encerrado"],           key=lambda x: x["dias"], reverse=True)
        indefinidos=        [l for l in todas_licitacoes if l["status"] == "Indefinido"]

        municipios_por_uf = {}
        for l in todas_licitacoes:
            uf = l["uf"]
            if uf not in municipios_por_uf:
                municipios_por_uf[uf] = set()
            municipios_por_uf[uf].add(l["municipio"])
        municipios_por_uf = {uf: sorted(cids) for uf, cids in sorted(municipios_por_uf.items())}

        global _ultima_busca
        _ultima_busca = {
            "termo":                    termo,
            "ufs_selecionadas":         ufs_selecionadas,
            "licitacoes":               todas_licitacoes,
            "abertos":                  abertos,
            "aguardando":               aguardando,
            "julgamento":               julgamento,
            "encerrados":               encerrados,
            "indefinidos":              indefinidos,
            "todos":                    abertos + aguardando + julgamento + encerrados + indefinidos,
            "data_inicial":             data_inicial or dt_ini.strftime("%Y-%m-%d"),
            "data_final":               data_final or dt_fim.strftime("%Y-%m-%d"),
            "incluir_sem_periodo":      incluir_sem_periodo,
            "total":                    len(todas_licitacoes),
            "municipios_por_uf":        municipios_por_uf,
            "limite_registros":         limite_registros,
            "modalidades":              MODALIDADES_DISPONIVEIS,
            "modalidades_selecionadas": modalidades_selecionadas,
        }

        print(f"[debug] Total final: {len(todas_licitacoes)}")
        print(f"[debug] Abertos: {len(abertos)}, Aguardando: {len(aguardando)}, Julgamento: {len(julgamento)}, Encerrados: {len(encerrados)}, Indefinidos: {len(indefinidos)}")
        if todas_licitacoes:
            print(f"[debug] Status ex: {todas_licitacoes[0].get('status')}")
            print(f"[debug] Objeto ex: {todas_licitacoes[0].get('objeto','')[:60]}")
        else:
            print("[debug] LISTA VAZIA!")
        yield f"data: {{\"tipo\": \"concluido\", \"total\": {len(todas_licitacoes)}, \"redirect\": \"/resultados\"}}\n\n"

    return StreamingResponse(
        gerar_eventos(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


# =====================================================
# ROTA: BUSCAR (POST normal — mantido para segurança)
# =====================================================

@app.post("/buscar", response_class=HTMLResponse)
async def buscar(request: Request):
    form = await request.form()
    termo = form.get("termo", "").strip()
    if not termo:
        return RedirectResponse("/", status_code=303)
    # Redireciona para o fluxo normal sem variações
    return RedirectResponse("/", status_code=303)


# =====================================================
# ROTA: RESULTADOS
# =====================================================

@app.get("/resultados", response_class=HTMLResponse)
async def resultados(request: Request):
    if not _ultima_busca:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="resultados.html",
        context={
            "estados": ESTADOS,
            "etps": etp_manager.listar_etps(),
            **_ultima_busca,
        },
    )


# =====================================================
# ROTA: EXPORTAR EXCEL
# =====================================================

@app.get("/exportar/excel")
async def exportar_excel():
    if not _ultima_busca:
        return RedirectResponse("/", status_code=303)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HTMLResponse("Instale openpyxl: pip install openpyxl", status_code=500)

    licitacoes = _ultima_busca.get("licitacoes", [])
    termo      = _ultima_busca.get("termo", "licitacoes")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licitações PNCP"

    cabecalhos = ["UF", "Município", "Status", "Prazo", "Encerramento",
                  "Publicação", "Objeto", "Valor Estimado", "Modalidade", "Órgão", "Link"]

    fill_header = PatternFill("solid", fgColor="1e3a8a")
    font_header = Font(color="FFFFFF", bold=True)
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=cab)
        cell.fill      = fill_header
        cell.font      = font_header
        cell.alignment = Alignment(horizontal="center")

    for row, l in enumerate(licitacoes, 2):
        ws.cell(row=row, column=1,  value=l["uf"])
        ws.cell(row=row, column=2,  value=l["municipio"])
        ws.cell(row=row, column=3,  value=l["status"])
        ws.cell(row=row, column=4,  value=l["prazo"])
        ws.cell(row=row, column=5,  value=l["encerramento"])
        ws.cell(row=row, column=6,  value=l["publicacao"])
        ws.cell(row=row, column=7,  value=l["objeto"])
        ws.cell(row=row, column=8,  value=l["valor"])
        ws.cell(row=row, column=9,  value=l["modalidade"])
        ws.cell(row=row, column=10, value=l["orgao"])
        ws.cell(row=row, column=11, value=l["link"])

    larguras = [6, 22, 12, 14, 20, 20, 70, 18, 18, 45, 55]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f"licitacoes_{termo.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
    )


# =====================================================
# INICIALIZAÇÃO LOCAL
# =====================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8001, reload=False)
